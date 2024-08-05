import pandas as pd
import streamlit as st
from info_search import infos
from groupy import grouping
from check_file import check
import io
import openpyxl as opxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
st.set_page_config(layout='wide')
unprocessed = []
def to_excel(df, HEADER=False, START=1):
    output = __import__("io").BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    pd.DataFrame(df).to_excel(writer, index=False, header=HEADER,startrow=START, startcol=START, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'})
    worksheet.set_column('A:A', None, format1)
    writer.close()
    return output.getvalue()
all_noms = []
st.session_state.uploaded_files = st.file_uploader('TECHNICAL SHEETS', accept_multiple_files=True)
p = 0
lp = len(st.session_state.uploaded_files)
if lp:
    pr_bar = st.progress(p, '')
exp = st.expander('Системы')
for file in st.session_state.uploaded_files:
    p+=1
    try:
        curinfo = infos(file)
        all_noms+=curinfo
        exp.markdown(f'<h2>Система {curinfo[0][0]}</h2>', unsafe_allow_html=True)
        for ci in curinfo:
            exp.markdown(f'<p><h4>{ci[1]}</h4>   <i>{ci[-1]} шт.</i></p>', unsafe_allow_html=True)
        exp.markdown('---')
    except Exception as e:
        e
        unprocessed.append(file.name)
    if lp:
        pr_bar.progress(p/lp, f"Обрабатывается {file.name}")
if lp:
    pr_bar.progress(p/lp, 'Готово')
if len(unprocessed):
    neob = st.expander('Необработанные файлы')
    for upcs in unprocessed:
        neob.markdown(f"<h4>{upcs}</h4>", unsafe_allow_html=True)
all_nomenclature_names = list(set(list(map(lambda x: x[1], all_noms))))
gnoms = grouping(all_noms)
tzinch = to_excel(list(map(lambda x: [x[1], x[2], x[0]], all_noms)))

if lp:
    try:
        st.expander('Таблица').table(gnoms)
        st.download_button(label='💾 Скачать файл для выгрузки в КП',data=to_excel(grouping(all_noms)), file_name= 'для кп.xls')
        st.download_button(label='💾 Скачать файл для выгрузки в КП (По системам)',data=to_excel(list(map(lambda x: [x[1], x[2], x[0]], all_noms))), file_name= 'для кп.xls')
        st.download_button(label='💾 Скачать проверочный файл',data=to_excel(pd.DataFrame(check(all_noms)),START=0) ,file_name= 'проверка.xlsx')
    except Exception as e:
        e
        pass
    try:
        list_oborud = {}
        upload_files = st.file_uploader("Загрузить КП из 1С", type=['xlsx', 'xls'])
        df1 = pd.read_excel(upload_files)[["Unnamed: 1","Unnamed: 3","Unnamed: 29", "Unnamed: 24", "Unnamed: 33", "Unnamed: 39"]]
        df2 = pd.read_excel(tzinch)[["Unnamed: 1","Unnamed: 2","Unnamed: 3"]]
        trans = {'K':'К'}
        serial_number = df1['Unnamed: 1'][9][25:38]
        df1 = df1[~ df1["Unnamed: 3"].isna()].reset_index(drop=True)
        for i in range(len(df1['Unnamed: 3'])):
            if df1['Unnamed: 3'][i].strip() != 'Товар':
                if 'КВАРК-ПН' in df1['Unnamed: 3'][i]:
                    df1['Unnamed: 3'][i] = df1['Unnamed: 3'][i].replace('Вентилятор ', '')
                if 'М5' in df1['Unnamed: 3'][i]:
                    df1['Unnamed: 3'][i] = df1['Unnamed: 3'][i].replace('-М5', '')
                if df1['Unnamed: 3'][i].strip() not in list_oborud.keys():
                    list_oborud[df1['Unnamed: 3'][i].strip().lower()] = [df1['Unnamed: 29'][i],df1['Unnamed: 33'][i]]
        df2['Unnamed: 4'] = ''
        df2['Unnamed: 5'] = ''
        for i in range(len(df2['Unnamed: 1'])):
            if 'ПНВ-40-20-20-2-380' in df2['Unnamed: 1'][i]:
                df2['Unnamed: 1'][i] = df2['Unnamed: 1'][i].translate(str.maketrans(trans))
            df2['Unnamed: 1'][i] = df2['Unnamed: 1'][i].translate(str.maketrans(trans))
            try:
                if df2['Unnamed: 1'][i][:-2].strip().lower() in list_oborud.keys():
                    df2['Unnamed: 1'][i] = df2['Unnamed: 1'][i][:-2]
                if df2['Unnamed: 1'][i].strip().lower() in list_oborud.keys():
                    df2['Unnamed: 4'][i] = list_oborud[df2['Unnamed: 1'][i].lower()][0]
                    df2['Unnamed: 5'][i] = round(df2['Unnamed: 4'][i]*df2['Unnamed: 2'][i],2)
            except:
                pass
        list_names_canal = list(df2['Unnamed: 1'])
        for y in range(len(list_names_canal)):
            list_names_canal[y] = list_names_canal[y].strip().lower()
        list_names_canal = set(list_names_canal)
        list_names_all = list(df1['Unnamed: 3'])
        for y in range(len(list_names_all)):
            list_names_all[y] = list_names_all[y].strip().lower()
        list_names_all = set(list_names_all)
        question_df = []
        final_df = []
        for i in range(len(df2['Unnamed: 1'])):
            if df2['Unnamed: 1'][i].strip().lower() in list_names_all:
                final_df.append([df2['Unnamed: 1'][i],df2['Unnamed: 2'][i],df2['Unnamed: 3'][i],df2['Unnamed: 4'][i],df2['Unnamed: 5'][i]])
            else:
                question_df.append([df2['Unnamed: 1'][i],df2['Unnamed: 2'][i],df2['Unnamed: 3'][i],df2['Unnamed: 4'][i],df2['Unnamed: 5'][i]])
        q = ''
        for i in range(len(df1['Unnamed: 3'])):
            if df1['Unnamed: 3'][i].strip() != 'Товар' and type(df1['Unnamed: 39'][i]) == str and df1['Unnamed: 3'][i].strip().lower() not in list_names_canal:
                df1['Unnamed: 39'][i] = list((df1['Unnamed: 39'][i].split(',')))

                for u in range(len(df1['Unnamed: 39'][i])):
                    if df1['Unnamed: 39'][i][u] != '' and df1['Unnamed: 39'][i][u][-1] == '.':
                        df1['Unnamed: 39'][i][u] = df1['Unnamed: 39'][i][u][:-1]
                        df1['Unnamed: 39'][i][u] = df1['Unnamed: 39'][i][u].strip()
                    elif df1['Unnamed: 39'][i][u] != '':
                        df1['Unnamed: 39'][i][u] = df1['Unnamed: 39'][i][u].strip()
                    else:
                        q = u
                if q != '':
                    del df1['Unnamed: 39'][i][q]
                    q = ''
                df1['Unnamed: 39'][i] = list(set(df1['Unnamed: 39'][i]))
                if len(df1['Unnamed: 39'][i]) == 1:
                    final_df.append([df1['Unnamed: 3'][i], df1['Unnamed: 24'][i], df1['Unnamed: 39'][i][0], df1['Unnamed: 29'][i], df1['Unnamed: 29'][i]*df1['Unnamed: 24'][i]])
                elif len(df1['Unnamed: 39'][i]) == df1['Unnamed: 24'][i]:
                    for k in range(len(df1['Unnamed: 39'][i])):
                        final_df.append([df1['Unnamed: 3'][i], 1, df1['Unnamed: 39'][i][k], df1['Unnamed: 29'][i], df1['Unnamed: 29'][i]])
                elif df1['Unnamed: 3'][i].strip().lower() not in list_names_canal:
                    question_df.append([df1['Unnamed: 3'][i], df1['Unnamed: 24'][i], df1['Unnamed: 39'][i], df1['Unnamed: 29'][i], df1['Unnamed: 29'][i]*df1['Unnamed: 24'][i]])
            elif df1['Unnamed: 3'][i].strip() != 'Товар' and type(df1['Unnamed: 39'][i]) != str:
                question_df.append([df1['Unnamed: 3'][i], df1['Unnamed: 24'][i], 'Система не указана', df1['Unnamed: 29'][i], df1['Unnamed: 29'][i]*df1['Unnamed: 24'][i]])
        final_df = pd.pd.DataFrame(final_df)
        final_df.drop_duplicates(subset=[0,1,2,3], inplace=True)
        final_df = final_df.sort_values(by=[2,0]).reset_index(drop=True)
        wb = opxl.Workbook()
        ws = wb.active
        thins = Side(border_style="thin", color="000000")
        system_info = ''
        counting_row = 16
        number_row = 1
        sum_price = 0
        total_amount = 0
        for i in range(len(final_df[0])):
            if i == 0:
                ws.column_dimensions['A'].width = 1.83
                ws.column_dimensions['A'].height = 15
                ws.column_dimensions['B'].width = 7
                ws.column_dimensions['D'].width = 7
                ws.column_dimensions['E'].width = 7
                ws.column_dimensions['C'].width = 55
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 20
                ws.row_dimensions[i+1].height = 30
                ws[f'B{i + counting_row}'].value = '№'
                ws[f'B{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'C{i + counting_row}'].value = 'Название'
                ws[f'C{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'C{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'C{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'D{i + counting_row}'].value = 'Кол-во'
                ws[f'D{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].alignment = Alignment(horizontal='center', wrap_text=True)
                ws[f'D{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'E{i + counting_row}'].value = 'Ед. изм.'
                ws[f'E{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'E{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'E{i + counting_row}'].alignment = Alignment(horizontal='center', wrap_text=True)
                ws[f'F{i + counting_row}'].value = 'Цена, с НДС'
                ws[f'F{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'F{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'F{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'G{i + counting_row}'].value = 'Сумма, с НДС'
                ws[f'G{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'G{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'G{i + counting_row}'].alignment = Alignment(horizontal='center')
                counting_row += 1
                ws[f'B{i + counting_row}'].value = f'Система ' + final_df[2][i].strip()
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'B{i + counting_row}:G{i + counting_row}')
                counting_row += 1
                ws[f'B{i + counting_row}'].value = number_row
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                number_row += 1
                ws[f'C{i + counting_row}'].value = final_df[0][i]
                ws[f'C{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].value = float(final_df[1][i])
                ws[f'D{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'D{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                ws[f'E{i + counting_row}'].value = 'шт.'
                ws[f'E{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'E{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'F{i + counting_row}'].value = final_df[3][i]
                ws[f'F{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'F{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'f{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[3][i] != 'nan' and final_df[3][i] != '':
                    pass
                else:
                    ws[f'f{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                ws[f'G{i + counting_row}'].value = final_df[4][i]
                ws[f'G{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'G{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'G{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[4][i] != 'nan' and final_df[4][i] != '':
                    sum_price += float(final_df[4][i])
                else:
                    ws[f'G{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                system_info = final_df[2][i].strip()
            elif system_info == '' or system_info != final_df[2][i].strip():
                counting_row += 1
                if system_info != final_df[2][i].strip() and system_info != '':
                    ws[f'B{i + counting_row - 1}'].value = 'Итого:'
                    ws[f'B{i + counting_row - 1}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'B{i + counting_row - 1}'].font = Font(size=12, bold=True)
                    ws[f'B{i + counting_row - 1}'].alignment = Alignment(horizontal='right')
                    ws.merge_cells(f'B{i + counting_row - 1}:F{i + counting_row - 1}')
                    ws[f'G{i + counting_row - 1}'].value = sum_price
                    ws[f'G{i + counting_row - 1}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'G{i + counting_row - 1}'].alignment = Alignment(horizontal='center')
                    ws[f'G{i + counting_row - 1}'].number_format = numbers.BUILTIN_FORMATS[4]
                    total_amount += sum_price
                    sum_price = 0
                    # counting_row += 1 (Дополнительный отступ между системами)
                ws[f'B{i + counting_row}'].value = f'Система ' + final_df[2][i]
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].font = Font(size=12, bold=True)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'B{i + counting_row}:G{i + counting_row}')
                counting_row += 1
                ws[f'B{i + counting_row}'].value = number_row
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                number_row += 1
                ws[f'C{i + counting_row}'].value = final_df[0][i]
                ws[f'C{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].value = float(final_df[1][i])
                ws[f'D{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'D{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                ws[f'E{i + counting_row}'].value = 'шт.'
                ws[f'E{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'E{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'F{i + counting_row}'].value = final_df[3][i]
                ws[f'F{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'F{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'f{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[3][i] != 'nan' and final_df[3][i] != '':
                    pass
                else:
                    ws[f'f{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                ws[f'G{i + counting_row}'].value = final_df[4][i]
                ws[f'G{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'G{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'G{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[4][i] != 'nan' and final_df[4][i] != '':
                    sum_price += float(final_df[4][i])
                else:
                    ws[f'G{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                system_info = final_df[2][i].strip()
            elif system_info == final_df[2][i].strip():
                ws[f'B{i + counting_row}'].value = number_row
                ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                number_row += 1
                ws[f'C{i + counting_row}'].value = final_df[0][i]
                ws[f'C{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].value = float(final_df[1][i])
                ws[f'D{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'D{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'D{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                ws[f'E{i + counting_row}'].value = 'шт.'
                ws[f'E{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'E{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'F{i + counting_row}'].value = final_df[3][i]
                ws[f'f{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[3][i] != 'nan' and final_df[3][i] != '':
                    pass
                else:
                    ws[f'f{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                ws[f'F{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'F{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'G{i + counting_row}'].value = final_df[4][i]
                ws[f'G{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                ws[f'G{i + counting_row}'].alignment = Alignment(horizontal='center')
                ws[f'G{i + counting_row}'].number_format = numbers.BUILTIN_FORMATS[4]
                if final_df[4][i] != 'nan' and final_df[4][i] != '':
                    sum_price += float(final_df[4][i])
                else:
                    ws[f'G{i + counting_row}'].fill = PatternFill('solid', fgColor="FF0000")
                if i == len(final_df[0])-1:
                    counting_row += 1
                    ws[f'B{i + counting_row }'].value = 'Итого:'
                    ws[f'B{i + counting_row }'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'B{i + counting_row }'].font = Font(size=12, bold=True)
                    ws[f'B{i + counting_row }'].alignment = Alignment(horizontal='right')
                    ws.merge_cells(f'B{i + counting_row }:F{i + counting_row }')
                    ws[f'G{i + counting_row }'].value = sum_price
                    ws[f'G{i + counting_row }'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'G{i + counting_row }'].alignment = Alignment(horizontal='center')
                    ws[f'G{i + counting_row }'].number_format = numbers.BUILTIN_FORMATS[4]
                    counting_row += 2
                    ws[f'B{i + counting_row }'].value = 'Общий Итог:'
                    ws[f'B{i + counting_row }'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'B{i + counting_row }'].font = Font(size=12, bold=True)
                    ws[f'B{i + counting_row }'].alignment = Alignment(horizontal='right')
                    ws.merge_cells(f'B{i + counting_row }:F{i + counting_row }')
                    total_amount += sum_price
                    ws[f'G{i + counting_row }'].value = total_amount
                    ws[f'G{i + counting_row }'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'G{i + counting_row }'].alignment = Alignment(horizontal='center')
                    ws[f'G{i + counting_row }'].number_format = numbers.BUILTIN_FORMATS[4]
                    counting_row += 2
                    ws[f'B{i + counting_row}'].value = '№'
                    ws[f'B{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'B{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'B{i + counting_row}'].alignment = Alignment(horizontal='center')
                    ws[f'C{i + counting_row}'].value = 'Название позиции которое не обработалось'
                    ws[f'C{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'C{i + counting_row}'].alignment = Alignment(horizontal='center')
                    ws[f'C{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'D{i + counting_row}'].value = 'Кол-во'
                    ws[f'D{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'D{i + counting_row}'].alignment = Alignment(horizontal='center', wrap_text=True)
                    ws[f'D{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'E{i + counting_row}'].value = 'Ед. изм.'
                    ws[f'E{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'E{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'E{i + counting_row}'].alignment = Alignment(horizontal='center', wrap_text=True)
                    ws[f'F{i + counting_row}'].value = 'Цена, с НДС'
                    ws[f'F{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'F{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'F{i + counting_row}'].alignment = Alignment(horizontal='center')
                    ws[f'G{i + counting_row}'].value = 'Системы'
                    ws[f'G{i + counting_row}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                    ws[f'G{i + counting_row}'].font = Font(size=12, bold=True)
                    ws[f'G{i + counting_row}'].alignment = Alignment(horizontal='center')
                    counting_row += 1
                    for l in range(len(question_df)):
                        ws[f'B{i + counting_row + l}'].value = number_row
                        ws[f'B{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        ws[f'B{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                        number_row += 1
                        ws[f'C{i + counting_row + l}'].value = question_df[l][0]
                        ws[f'C{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        ws[f'D{i + counting_row + l}'].value = float(question_df[l][1])
                        ws[f'D{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        ws[f'D{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                        ws[f'D{i + counting_row + l}'].number_format = numbers.BUILTIN_FORMATS[4]
                        ws[f'E{i + counting_row + l}'].value = 'шт.'
                        ws[f'E{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        ws[f'E{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                        ws[f'F{i + counting_row + l}'].value = question_df[l][3]
                        ws[f'F{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                        ws[f'F{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                        ws[f'f{i + counting_row + l}'].number_format = numbers.BUILTIN_FORMATS[4]

                        if type(question_df[l][2]) == str:
                            ws[f'G{i + counting_row + l}'].value = question_df[l][2]
                            ws[f'G{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                            ws[f'G{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                            ws[f'G{i + counting_row + l}'].number_format = numbers.BUILTIN_FORMATS[4]
                        else:
                            ws[f'G{i + counting_row + l}'].value = ' '.join(question_df[l][2])
                            ws[f'G{i + counting_row + l}'].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                            ws[f'G{i + counting_row + l}'].alignment = Alignment(horizontal='center')
                            ws[f'G{i + counting_row + l}'].number_format = numbers.BUILTIN_FORMATS[4]
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        download = st.download_button(label='📥 Скачать', data=virtual_workbook , file_name=f'Кп {serial_number} Каналка.xlsx')
    except Exception as error:
        # st.write(error)
        pass
